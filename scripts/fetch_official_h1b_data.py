#!/usr/bin/env python3
import argparse
import csv
import pathlib
import re
import subprocess
from openpyxl import load_workbook

DOL_LCA_XLSX_URL = "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/LCA_Disclosure_Data_FY2026_Q1.xlsx"
USCIS_CSV_URL = "https://www.uscis.gov/sites/default/files/document/data/h1b_datahubexport-2023.csv"


def download_file(url: str, target: pathlib.Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            "curl",
            "-L",
            "--fail",
            url,
            "-o",
            str(target),
        ],
        check=True,
    )


def as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value):
    text = as_text(value)
    if not text:
        return ""
    normalized = re.sub(r"[$,\s]", "", text)
    try:
        return float(normalized)
    except ValueError:
        return ""


def parse_year(row: dict) -> int:
    candidates = [
        row.get("CASE_SUBMITTED"),
        row.get("RECEIVED_DATE"),
        row.get("DECISION_DATE"),
        row.get("BEGIN_DATE"),
        row.get("END_DATE"),
        row.get("CASE_RECEIVED_DATE"),
    ]

    for value in candidates:
        text = as_text(value)
        match = re.search(r"(20\d{2})", text)
        if match:
            return int(match.group(1))

    return 2026


def convert_dol_xlsx_to_normalized_csv(
    source_xlsx: pathlib.Path,
    output_csv: pathlib.Path,
    max_rows: int | None,
) -> int:
    workbook = load_workbook(filename=source_xlsx,
                             read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]

    rows_iter = worksheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise RuntimeError("DOL workbook has no header row.")

    headers = [as_text(value) for value in header]

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    count = 0

    with output_csv.open("w", newline="", encoding="utf-8") as output_handle:
        writer = csv.writer(output_handle)
        writer.writerow(["employer", "job_title", "country",
                        "work_location", "wage", "status", "year"])

        for row_values in rows_iter:
            row = {headers[index]: row_values[index]
                   for index in range(min(len(headers), len(row_values)))}

            visa_class = as_text(row.get("VISA_CLASS")).upper()
            if visa_class and visa_class not in {"H-1B", "H-1B1", "E-3"}:
                continue

            employer = as_text(row.get("EMPLOYER_NAME")
                               or row.get("EMPLOYER_NAME_DECLARED"))
            job_title = as_text(row.get("JOB_TITLE") or row.get("SOC_TITLE"))
            country = as_text(row.get("WORKSITE_COUNTRY") or row.get(
                "COUNTRY_OF_CITIZENSHIP") or "Unknown")
            city = as_text(row.get("WORKSITE_CITY"))
            state = as_text(row.get("WORKSITE_STATE"))
            work_location = ", ".join([part for part in [city, state] if part])
            wage = (
                parse_number(row.get("WAGE_RATE_OF_PAY_FROM"))
                or parse_number(row.get("WAGE_RATE_OF_PAY_TO"))
                or parse_number(row.get("PREVAILING_WAGE"))
            )
            status = as_text(row.get("CASE_STATUS") or row.get("STATUS"))
            year = parse_year(row)

            writer.writerow([employer, job_title, country,
                            work_location, wage, status, year])
            count += 1

            if max_rows is not None and count >= max_rows:
                break

    workbook.close()
    return count


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--max-rows", type=int, default=None)
    args = parser.parse_args()

    root = pathlib.Path(__file__).resolve().parents[1]
    data_dir = root / "apps" / "web" / "public" / "data"

    dol_xlsx_path = data_dir / "LCA_Disclosure_Data_FY2026_Q1.xlsx"
    dol_csv_path = data_dir / "dol_lca_h1b_fy2026_q1.csv"
    uscis_csv_path = data_dir / "uscis_h1b_employer_data_hub_2023.csv"

    print("Downloading USCIS H-1B Employer Data Hub CSV...")
    download_file(USCIS_CSV_URL, uscis_csv_path)

    print("Downloading DOL LCA disclosure XLSX...")
    download_file(DOL_LCA_XLSX_URL, dol_xlsx_path)

    print("Converting DOL XLSX to normalized CSV...")
    normalized_count = convert_dol_xlsx_to_normalized_csv(
        dol_xlsx_path, dol_csv_path, args.max_rows)

    if dol_xlsx_path.exists():
        dol_xlsx_path.unlink()

    print("Done.")
    print(f"USCIS CSV: {uscis_csv_path}")
    print(f"DOL XLSX (removed after conversion): {dol_xlsx_path}")
    print(f"DOL normalized CSV rows: {normalized_count}")
    print(f"DOL normalized CSV: {dol_csv_path}")


if __name__ == "__main__":
    main()
